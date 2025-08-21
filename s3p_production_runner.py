#!/usr/bin/env python3
"""
S3P Production API Runner
A streamlined Python script for production testing of core S3P API services:
- Cashin (MTN Mobile Money deposits)
- Cashout (MTN Mobile Money withdrawals) 
- Topup (Orange airtime purchases)
"""

import requests
import hashlib
import hmac
import time
import base64
import json
import argparse
import os
import sys
from urllib.parse import quote
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
from enum import Enum
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

class ServiceType(Enum):
    """Supported production service types"""
    CASHIN = "cashin"
    CASHOUT = "cashout"
    TOPUP = "topup"

class TransactionStatus(Enum):
    """Transaction status values"""
    PENDING = "PENDING"
    SUCCESS = "SUCCESS"
    FAILED = "FAILED"
    ERROR = "ERROR"

@dataclass
class TransactionConfig:
    """Configuration for a single transaction test"""
    service_type: ServiceType
    service_id: str
    amount: int
    customer_phone: str
    customer_email: str
    customer_name: str
    customer_address: str
    service_number: str
    transaction_id: str

@dataclass
class TransactionResult:
    """Result of a transaction test"""
    config: TransactionConfig
    success: bool
    payment_item_id: Optional[str] = None
    quote_id: Optional[str] = None
    ptn: Optional[str] = None
    final_status: Optional[str] = None
    error_message: Optional[str] = None
    execution_time: Optional[float] = None

class HmacService:
    """HMAC authentication service for S3P API"""
    
    def __init__(self, s3p_url: str, s3p_key: str, s3p_secret: str):
        self.s3p_url = s3p_url
        self.s3p_key = s3p_key
        self.s3p_secret = s3p_secret

    def generate_auth_header(self, http_method: str, query_params: Optional[Dict] = None, 
                           request_data: Optional[Dict] = None) -> str:
        """Generate HMAC-SHA1 authentication header for S3P API"""
        if query_params is None:
            query_params = {}
        if request_data is None:
            request_data = {}

        # Generate timestamp and nonce
        timestamp = str(int(round(time.time() * 1000)))
        nonce = str(int(round(time.time() * 1000)))

        # Construct the s3pParams dictionary
        s3p_params = {
            "s3pAuth_nonce": nonce,
            "s3pAuth_timestamp": timestamp,
            "s3pAuth_signature_method": "HMAC-SHA1",
            "s3pAuth_token": self.s3p_key
        }

        # Combine all parameters
        input_data = {**query_params, **request_data}
        params = {**input_data, **s3p_params}

        # Clean parameter values
        params = {k: (v.strip() if isinstance(v, str) else v) for k, v in params.items()}

        # Sort parameters alphabetically
        sorted_params = sorted(params.items())

        # Construct parameter string
        parameter_string = "&".join(f"{key}={str(value)}" for key, value in sorted_params)

        # Construct base string
        base_string = f"{http_method.upper()}&{quote(self.s3p_url, safe='-')}&{quote(parameter_string, safe='-')}"

        # Generate signature
        hashed = hmac.new(bytes(self.s3p_secret, "utf-8"), bytes(base_string, "utf-8"), hashlib.sha1)
        signature = base64.b64encode(hashed.digest()).decode('utf-8')

        # Construct auth header
        auth_header = (
            f's3pAuth '
            f's3pAuth_timestamp="{timestamp}", '
            f's3pAuth_nonce="{nonce}", '
            f's3pAuth_signature_method="HMAC-SHA1", '
            f's3pAuth_token="{self.s3p_key}", '
            f's3pAuth_signature="{signature}"'
        )

        return auth_header

class S3PApiClient:
    """S3P API client with authentication"""
    
    def __init__(self, base_url: str, api_key: str, api_secret: str):
        self.base_url = base_url.rstrip('/')
        self.api_key = api_key
        self.api_secret = api_secret
        self.session = requests.Session()
        self.session.headers.update({'Content-Type': 'application/json'})

    def _make_request(self, method: str, endpoint: str, query_params: Optional[Dict] = None, 
                     request_data: Optional[Dict] = None) -> requests.Response:
        """Make authenticated request to S3P API"""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        hmac_service = HmacService(url, self.api_key, self.api_secret)
        
        auth_header = hmac_service.generate_auth_header(method, query_params, request_data)
        headers = {'Authorization': auth_header}
        
        if method.upper() == 'GET':
            response = self.session.get(url, headers=headers, params=query_params or {})
        elif method.upper() == 'POST':
            response = self.session.post(url, headers=headers, json=request_data or {})
        else:
            raise ValueError(f"Unsupported HTTP method: {method}")
        
        return response

    def get_payment_items(self, service_type: ServiceType, service_id: str) -> Dict[str, Any]:
        """Get payment items for a service"""
        endpoint = service_type.value
        query_params = {"serviceid": service_id}
        
        response = self._make_request("GET", endpoint, query_params=query_params)
        response.raise_for_status()
        return response.json()

    def get_quote(self, payment_item_id: str, amount: int) -> Dict[str, Any]:
        """Get quote for a payment item"""
        request_data = {
            "payItemId": payment_item_id,
            "amount": amount
        }
        
        response = self._make_request("POST", "quotestd", request_data=request_data)
        response.raise_for_status()
        return response.json()

    def execute_payment(self, quote_id: str, customer_phone: str, customer_email: str,
                       customer_name: str, customer_address: str, service_number: str,
                       transaction_id: str) -> Dict[str, Any]:
        """Execute payment using quote"""
        request_data = {
            "quoteId": quote_id,
            "customerPhonenumber": customer_phone,
            "customerEmailaddress": customer_email,
            "customerName": customer_name,
            "customerAddress": customer_address,
            "serviceNumber": service_number,
            "trid": transaction_id
        }
        
        response = self._make_request("POST", "collectstd", request_data=request_data)
        response.raise_for_status()
        return response.json()

    def verify_transaction(self, ptn: str) -> Dict[str, Any]:
        """Verify transaction status"""
        query_params = {"ptn": ptn}
        
        response = self._make_request("GET", "verifytx", query_params=query_params)
        response.raise_for_status()
        return response.json()

class S3PProductionRunner:
    """Production test runner for core S3P API services"""
    
    def __init__(self, api_client: S3PApiClient, verbose: bool = False):
        self.api_client = api_client
        self.verbose = verbose

    def print_status(self, message: str, level: str = "INFO"):
        """Print status message with timestamp"""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        colors = {
            "INFO": "",
            "SUCCESS": "\033[92m",  # Green
            "WARNING": "\033[93m",  # Yellow
            "ERROR": "\033[91m"     # Red
        }
        reset = "\033[0m"
        color = colors.get(level, "")
        print(f"{color}[{timestamp}] {level}: {message}{reset}")

    def execute_transaction(self, config: TransactionConfig) -> TransactionResult:
        """Execute a single transaction with the 4-step S3P flow"""
        result = TransactionResult(config, False)
        start_time = time.time()
        
        self.print_status(f"Starting {config.service_type.value} transaction {config.transaction_id}")
        
        try:
            # Step 1: Get payment items
            self.print_status(f"Step 1: Getting payment items for service {config.service_id}")
            payment_items_response = self.api_client.get_payment_items(
                config.service_type, config.service_id
            )
            
            if self.verbose:
                self.print_status(f"Payment items response: {json.dumps(payment_items_response, indent=2)}")
            
            # Extract payment item ID
            if isinstance(payment_items_response, list) and payment_items_response:
                payment_item = payment_items_response[0]
                payment_item_id = payment_item.get('payItemId')
            else:
                payment_item_id = payment_items_response.get('payItemId')
                
            if not payment_item_id:
                raise Exception("No payment item ID found in response")
            
            result.payment_item_id = payment_item_id
            self.print_status(f"Found payment item ID: {payment_item_id}")
            
            # Step 2: Get quote
            self.print_status(f"Step 2: Getting quote for amount {config.amount}")
            quote_response = self.api_client.get_quote(payment_item_id, config.amount)
            
            if self.verbose:
                self.print_status(f"Quote response: {json.dumps(quote_response, indent=2)}")
            
            quote_id = quote_response.get('quoteId')
            if not quote_id:
                raise Exception("No quote ID found in response")
            
            result.quote_id = quote_id
            self.print_status(f"Received quote ID: {quote_id}")
            
            # Step 3: Execute payment
            self.print_status(f"Step 3: Executing payment")
            payment_response = self.api_client.execute_payment(
                quote_id, config.customer_phone, config.customer_email,
                config.customer_name, config.customer_address, 
                config.service_number, config.transaction_id
            )
            
            if self.verbose:
                self.print_status(f"Payment response: {json.dumps(payment_response, indent=2)}")
            
            ptn = payment_response.get('ptn')
            if not ptn:
                raise Exception("No PTN found in payment response")
            
            result.ptn = ptn
            self.print_status(f"Payment initiated with PTN: {ptn}")
            
            # Step 4: Wait and verify
            self.print_status(f"Step 4: Waiting 20 seconds before verification")
            time.sleep(20)
            
            self.print_status("Verifying transaction status")
            verify_response = self.api_client.verify_transaction(ptn)
            
            if self.verbose:
                self.print_status(f"Verification response: {json.dumps(verify_response, indent=2)}")
            
            # Extract final status from polling - handle array response
            verification_data = None
            if isinstance(verify_response, list) and verify_response:
                verification_data = verify_response[0]
            elif isinstance(verify_response, dict):
                verification_data = verify_response
            
            if verification_data and 'status' in verification_data:
                final_status = verification_data['status']
                result.final_status = final_status
                
                if final_status == TransactionStatus.SUCCESS.value:
                    result.success = True
                    self.print_status(f"Transaction {config.transaction_id} completed successfully!", "SUCCESS")
                elif final_status == TransactionStatus.PENDING.value:
                    self.print_status(f"Transaction {config.transaction_id} is still pending", "WARNING")
                else:
                    self.print_status(f"Transaction {config.transaction_id} failed with status: {final_status}", "ERROR")
            else:
                raise Exception("No status found in verification response")

        except requests.exceptions.RequestException as e:
            error_msg = f"API request failed: {str(e)}"
            result.error_message = error_msg
            self.print_status(error_msg, "ERROR")
        except Exception as e:
            error_msg = f"Unknown error: {str(e)}"
            result.error_message = error_msg
            self.print_status(error_msg, "ERROR")
        finally:
            result.execution_time = time.time() - start_time
        
        return result

    def run_transactions(self, configs: List[TransactionConfig]) -> List[TransactionResult]:
        """Execute multiple transactions"""
        results = []
        total_transactions = len(configs)
        
        self.print_status(f"Starting execution of {total_transactions} transactions")
        
        for i, config in enumerate(configs, 1):
            self.print_status(f"\n=== Transaction {i}/{total_transactions} ===")
            result = self.execute_transaction(config)
            results.append(result)
            
            # Add delay between transactions
            if i < total_transactions:
                time.sleep(2)
        
        # Print summary
        successful = sum(1 for r in results if r.success)
        failed = total_transactions - successful
        
        self.print_status(f"\n=== EXECUTION SUMMARY ===")
        self.print_status(f"Total transactions: {total_transactions}")
        self.print_status(f"Successful: {successful}", "SUCCESS" if successful > 0 else "INFO")
        self.print_status(f"Failed: {failed}", "ERROR" if failed > 0 else "INFO")
        
        if failed > 0:
            self.print_status(f"\nFailed transactions:")
            for result in results:
                if not result.success:
                    self.print_status(f"  - {result.config.transaction_id}: {result.error_message}", "ERROR")
        
        # Generate Excel report
        report_filename = self.generate_excel_report(results)
        self.print_status(f"\nExcel report generated: {report_filename}", "SUCCESS")
        
        return results

    def generate_excel_report(self, results: List[TransactionResult]) -> str:
        """Generate Excel report with transaction results"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "S3P Production Test Results"
        
        # Headers
        headers = ["Service ID", "Amount", "Service Type", "Service Number", "PTN", "Transaction ID", "Status", "Error Message"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        for row, result in enumerate(results, 2):
            worksheet.cell(row=row, column=1, value=result.config.service_id)
            worksheet.cell(row=row, column=2, value=result.config.amount)
            worksheet.cell(row=row, column=3, value=result.config.service_type.value)
            worksheet.cell(row=row, column=4, value=result.config.service_number)
            worksheet.cell(row=row, column=5, value=result.ptn or "Not Generated")
            worksheet.cell(row=row, column=6, value=result.config.transaction_id)
            worksheet.cell(row=row, column=7, value="SUCCESS" if result.success else "FAILED")
            worksheet.cell(row=row, column=8, value=result.error_message or "")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"s3p_production_report_{timestamp}.xlsx"
        workbook.save(filename)
        
        return filename

def generate_unique_transaction_id(service_type: str) -> str:
    """Generate unique transaction ID starting with JAY and at least 21 characters"""
    import random
    timestamp = int(time.time())
    random_suffix = f"{random.randint(1000, 9999)}"
    # Format: JAY + service_type + timestamp + random_suffix (guaranteed 21+ chars)
    transaction_id = f"JAY{service_type.upper()}{timestamp}{random_suffix}"
    return transaction_id

# Production service configurations
PRODUCTION_SERVICES = {
    ServiceType.CASHIN: {
        "service_id": "20052",
        "service_number": "677389120",
        "default_amount": 1000
    },
    ServiceType.CASHOUT: {
        "service_id": "20053", 
        "service_number": "677389120",
        "default_amount": 500
    },
    ServiceType.TOPUP: {
        "service_id": "20062",
        "service_number": "698081976",
        "default_amount": 500
    }
}

# Default customer data for production testing
DEFAULT_CUSTOMER = {
    "phone": "237655754334",
    "email": "test@smobilpay.com", 
    "name": "Test Customer",
    "address": "Test Address, Douala, Cameroon"
}

def create_default_configs() -> List[TransactionConfig]:
    """Create default test configurations for all production services"""
    configs = []
    
    for service_type, service_config in PRODUCTION_SERVICES.items():
        config = TransactionConfig(
            service_type=service_type,
            service_id=service_config["service_id"],
            amount=service_config["default_amount"],
            customer_phone=DEFAULT_CUSTOMER["phone"],
            customer_email=DEFAULT_CUSTOMER["email"],
            customer_name=DEFAULT_CUSTOMER["name"],
            customer_address=DEFAULT_CUSTOMER["address"],
            service_number=service_config["service_number"],
            transaction_id=generate_unique_transaction_id(service_type.value)
        )
        configs.append(config)
    
    return configs

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description="S3P Production API Runner - Test core money transfer services",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python s3p_production_runner.py --all
  python s3p_production_runner.py --single cashin 1000
  python s3p_production_runner.py --single topup 500 --verbose
        """
    )
    
    # API credentials
    parser.add_argument('--url', default=os.getenv('S3P_URL', 'https://s3p.smobilpay.staging.maviance.info/v2'),
                       help='S3P API base URL')
    parser.add_argument('--key', default=os.getenv('S3P_KEY', ''),
                       help='S3P API key')
    parser.add_argument('--secret', default=os.getenv('S3P_SECRET', ''),
                       help='S3P API secret')
    
    # Execution modes
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument('--all', action='store_true',
                           help='Test all production services (cashin, cashout, topup)')
    mode_group.add_argument('--single', nargs=2, metavar=('SERVICE_TYPE', 'AMOUNT'),
                           help='Test single service (cashin|cashout|topup amount)')
    
    # Options
    parser.add_argument('--verbose', action='store_true',
                       help='Enable verbose output with API request/response details')
    
    args = parser.parse_args()
    
    # Validate credentials
    if not args.key or not args.secret:
        print("Error: S3P API credentials are required.")
        print("Set S3P_KEY and S3P_SECRET environment variables or use --key and --secret arguments")
        sys.exit(1)
    
    # Create API client
    api_client = S3PApiClient(args.url, args.key, args.secret)
    runner = S3PProductionRunner(api_client, verbose=args.verbose)
    
    # Determine configurations to run
    if args.all:
        configs = create_default_configs()
    else:
        service_type_str, amount_str = args.single
        
        try:
            service_type = ServiceType(service_type_str.lower())
            amount = int(amount_str)
        except (ValueError, KeyError):
            print(f"Error: Invalid service type '{service_type_str}'. Use: cashin, cashout, or topup")
            sys.exit(1)
        
        if service_type not in PRODUCTION_SERVICES:
            print(f"Error: Service type '{service_type_str}' not supported in production runner")
            sys.exit(1)
        
        service_config = PRODUCTION_SERVICES[service_type]
        config = TransactionConfig(
            service_type=service_type,
            service_id=service_config["service_id"],
            amount=amount,
            customer_phone=DEFAULT_CUSTOMER["phone"],
            customer_email=DEFAULT_CUSTOMER["email"],
            customer_name=DEFAULT_CUSTOMER["name"],
            customer_address=DEFAULT_CUSTOMER["address"],
            service_number=service_config["service_number"],
            transaction_id=generate_unique_transaction_id(service_type.value)
        )
        configs = [config]
    
    # Execute transactions
    try:
        results = runner.run_transactions(configs)
        
        # Exit with appropriate code
        failed_count = sum(1 for r in results if not r.success)
        sys.exit(0 if failed_count == 0 else 1)
        
    except KeyboardInterrupt:
        print("\nTest execution interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()